import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
from mappings import mapping_ru, mapping_en  # обязательно имя файла mappings.py

def transform_excel(uploaded_file, language="ru"):
    try:
        # Выбор шаблона по языку
        template_name = "ТР.xlsx" if language == "ru" else "ТР EN.xlsx"
        template_path = os.path.join(os.path.dirname(__file__), template_name)

        # Чтение исходного файла
        df_raw = pd.read_excel(uploaded_file, header=7).reset_index(drop=True)

        # Функция для умного преобразования значений
        def smart_convert(value):
            if pd.isna(value):
                return value
            if isinstance(value, str):
                value = value.strip().replace(',', '.')
                try:
                    return float(value)
                except ValueError:
                    return value
            return value

        # Приведение чисел к числовому типу
        for col_idx in [21, 0, 26, 27, 13, 14, 3, 33, 35, 16, 20, 19, 18, 20, 38, 39, 37, 40, 34, 36]:
            df_raw.iloc[:, col_idx] = df_raw.iloc[:, col_idx].apply(smart_convert)

        # Структура выходного файла
        columns = [
            "№ линии", "НАИМЕНОВАНИЕ ТРУБОПРОВОДА", "ДИАМЕТР (мм)", "ДЛИНА ТРУБОПРОВОДА (М)",
            "ТЕМПЕРАТУРА ПОДДЕРЖАНИЯ (°C)", "МАКСИМАЛЬНАЯ ТЕХНОЛОГИЧЕСКАЯ ТЕМПЕРАТУРА ПРОЦЕССА (°C)",
            "МИНИМАЛЬНАЯ ТЕМПЕРАТУРА ОКРУЖАЮЩЕЙ СРЕДЫ (°C)", "МАКСИМАЛЬНАЯ ТЕМПЕРАТУРА ОКРУЖАЮЩЕЙ СРЕДЫ",
            "Т-КЛАСС", "ТЕМПЕРАТУРА ПУСКА (°C)", "ПРОПАРКА (°C)", "ТОЛЩИНА ТЕПЛОИЗОЛЯЦИИ (мм)",
            "ЗАДВИЖКИ (шт.)", "ФЛАНЦЕВ (шт.)", "ОПОРЫ (шт.)", "УДЕЛЬНЫЕ ТЕПЛОПОТЕРИ ТРУБОПРОВОДА (Вт/м)",
            "МЕТОД РЕГУЛИРОВАНИЯ ТЕМПЕРАТУРЫ", "НАПРЯЖЕНИЕ ПИТАНИЯ (В)", "РАСЧЕТНАЯ МОЩНОСТЬ (Вт)",
            "РАБОЧИЙ ТОК (А)", "ДЛИТЕЛЬНЫЙ ПУСКОВОЙ ТОК (А)", "КОЛИЧЕСТВО НАБОРОВ", "ФАКТОР ПРОКЛАДКИ",
            "УДЕЛЬНАЯ МОЩНОСТЬ КАБЕЛЯ (Вт/м)", "МАКСИМАЛЬНАЯ ТЕМПЕРАТУРА ОБОЛОЧКИ КАБЕЛЯ (°C)",
            "КОЛИЧЕСТВО НАГРЕВАТЕЛЬНОГО КАБЕЛЯ (м)", "ТИП НАГРЕВАТЕЛЬНОГО КАБЕЛЯ"
        ]
        df_result = pd.DataFrame(columns=columns)

        # Заполнение значений
        df_result["№ линии"] = df_raw.iloc[:, 21]
        df_result["НАИМЕНОВАНИЕ ТРУБОПРОВОДА"] = df_raw.iloc[:, 0]
        df_result["ДИАМЕТР (мм)"] = df_raw.iloc[:, 26]
        df_result["ДЛИНА ТРУБОПРОВОДА (М)"] = df_raw.iloc[:, 27]
        df_result["ТЕМПЕРАТУРА ПОДДЕРЖАНИЯ (°C)"] = df_raw.iloc[:, 13]
        df_result["МАКСИМАЛЬНАЯ ТЕХНОЛОГИЧЕСКАЯ ТЕМПЕРАТУРА ПРОЦЕССА (°C)"] = df_raw.iloc[:, 14]
        df_result["Т-КЛАСС"] = df_raw.iloc[:, 3]
        df_result["ТОЛЩИНА ТЕПЛОИЗОЛЯЦИИ (мм)"] = df_raw.iloc[:, 33]
        df_result["УДЕЛЬНЫЕ ТЕПЛОПОТЕРИ ТРУБОПРОВОДА (Вт/м)"] = df_raw.iloc[:, 35]
        df_result["МЕТОД РЕГУЛИРОВАНИЯ ТЕМПЕРАТУРЫ"] = df_raw.iloc[:, 16]
        df_result["НАПРЯЖЕНИЕ ПИТАНИЯ (В)"] = df_raw.iloc[:, 20]
        df_result["РАСЧЕТНАЯ МОЩНОСТЬ (Вт)"] = df_raw.iloc[:, 19]
        df_result["РАБОЧИЙ ТОК (А)"] = df_raw.iloc[:, 18]
        df_result["ДЛИТЕЛЬНЫЙ ПУСКОВОЙ ТОК (А)"] = df_raw.iloc[:, 20]
        df_result["КОЛИЧЕСТВО НАБОРОВ"] = df_raw.iloc[:, 38]
        df_result["ФАКТОР ПРОКЛАДКИ"] = df_raw.iloc[:, 39]
        df_result["УДЕЛЬНАЯ МОЩНОСТЬ КАБЕЛЯ (Вт/м)"] = df_raw.iloc[:, 37]
        df_result["МАКСИМАЛЬНАЯ ТЕМПЕРАТУРА ОБОЛОЧКИ КАБЕЛЯ (°C)"] = df_raw.iloc[:, 40]
        df_result["КОЛИЧЕСТВО НАГРЕВАТЕЛЬНОГО КАБЕЛЯ (м)"] = df_raw.iloc[:, 34]

        # Перевод марки кабеля
        mapping = mapping_ru if language == "ru" else mapping_en
        df_result["ТИП НАГРЕВАТЕЛЬНОГО КАБЕЛЯ"] = df_raw.iloc[:, 36].apply(
            lambda x: mapping.get(x, x) if isinstance(x, str) else x
        )

        # Загрузка шаблона и вставка данных
        wb = load_workbook(template_path)
        ws = wb.active
        data_font = Font(name="Times New Roman", size=12)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        for row_idx, row in df_result.iterrows():
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx + 2, column=col_idx, value=None if pd.isna(value) else value)
                cell.font = data_font
                cell.border = thin_border

        now = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
        output_filename = f"Базовый расчет {now}.xlsx"
        output_path = os.path.join(os.path.dirname(__file__), output_filename)
        wb.save(output_path)

        return output_path

    except Exception as e:
        return f"Ошибка при обработке файла: {e}"
